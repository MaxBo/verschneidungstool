# -*- coding: utf-8 -*-
from typing import List
import psycopg2
from verschneidungstool.connection import Connection
from verschneidungstool.config import Config
from verschneidungstool.save2visumtransfer import save_to_visum_transfer
from PyQt5 import QtCore
import tempfile, os, shutil, re
import csv
from openpyxl import load_workbook, Workbook
import pandas as pd

config = Config()


class DBConnection(object):
    def __init__(self, login):
        self.login = login
        self.colums_available = None
        self.vt_schema = 'verschneidungstool'

    def fetch(self, sql):
        with Connection(login=self.login) as conn:
            self.conn = conn
            cursor = self.conn.cursor()
            cursor.execute(sql)
            rows = cursor.fetchall()
        return rows

    def copy_expert(self, sql, fileobject):
        with Connection(login=self.login) as conn:
            self.conn = conn
            cursor = self.conn.cursor()
            cursor.copy_expert(sql, fileobject)

    def execute(self, sql):
        with Connection(login=self.login) as conn:
            self.conn = conn
            cursor = self.conn.cursor()
            try:
                cursor.execute(sql)
            except psycopg2.ProgrammingError as e:
                raise psycopg2.ProgrammingError(
                    os.linesep.join((sql, str(e))))
            conn.commit()

    def get_areas_available(self):
        sql = f"""
        SELECT id, area_name, schema, table_name, can_be_deleted,
        check_last_calculation
        FROM {self.vt_schema}.areas_available
        ORDER BY id
        """
        return self.fetch(sql)

    def get_scenarios_available(self):
        sql = f"""
        SELECT scenario
        FROM {self.vt_schema}.scenarios_available
        ORDER BY jahr, scenario
        """
        return self.fetch(sql)

    def get_year_of_scenario(self, scenario: str) -> str:
        sql = f"""
        SELECT jahr AS year
        FROM {self.vt_schema}.scenarios_available sa
        WHERE sa.scenario = '{scenario}'
        """
        return self.fetch(sql)

    def get_resulttables_available(self):
        sql = f"""
        SELECT schema_table, visum_class, long_format
        FROM {self.vt_schema}.resulttables_available
        ORDER BY schema_table
        """
        return self.fetch(sql)

    def get_tables_to_download(self):
        sql = f"""
        SELECT id, name, schema, tablename, category
        FROM {self.vt_schema}.tables_to_download
        ORDER BY category
        """
        return self.fetch(sql)

    def get_structure_available(self, year):
        if not self.colums_available:
            sql_col_avail = f"""
            SELECT *
            FROM {self.vt_schema}.columns_available
            ORDER BY table_type, id
            """
            col_avail = self.fetch(sql_col_avail)
            self.colums_available = {}
            for record in col_avail:
                if record.table_type not in self.colums_available:
                    self.colums_available[record.table_type] = []
                self.colums_available[record.table_type].append({
                    'name': record.column,
                    'description': record.description,
                    'resulttable': record.resulttable,
                })

        sql_cat = f"""
        SELECT tc.name
        FROM {self.vt_schema}.table_categories AS tc,
        {self.vt_schema}.column_definitions AS cd
        WHERE tc.name = cd.table_category
        AND cd.from_year <= {year}
        AND cd.to_year >= {year}
        ORDER BY id
        """
        categories = self.fetch(sql_cat)
        structure = {}

        for record in categories:
            structure[record.name] = self.colums_available.get(record.name, dict())
        return structure

    def get_column_definition(self, colname: str, parent: str):
        """
        get column definition for colname
        """
        coldefs = self.colums_available[parent]
        for coldef in coldefs:
            if coldef['name'] == colname:
                return coldef

    def get_schemata_available(self):
        sql = f"""
        SELECT name
        FROM {self.vt_schema}.schemata_available
        """
        return self.fetch(sql)

    def drop_area(self, id, table, schema):

        last = self.get_last_calculated()
        if schema == last[0].schema and table == last[0].table_name:
            return False,
        (f'{schema}.{table} kann nicht gelöscht werden,\n'
         'da die letzte Verschneidung mit dieser Aggregation erfolgte. \n\n'
         'Bitte führen Sie zunächst eine Verschneidung mit einer\n'
         'anderen Aggregationsstufe durch, bevor Sie diese löschen.')
        # remove row from available schemata
        sql_remove = f"""
        DELETE FROM {self.vt_schema}.areas_available
        WHERE id = {id}
        """
        # drop table
        sql_drop = f"""
        Drop TABLE IF EXISTS "{schema}"."{table}"
        """
        try:
            self.execute(sql_remove)
            self.execute(sql_drop)
            return True, f'Löschen von {schema}.{table} erfolgreich'
        except:
            return False, ('Ein datenbankinterner Fehler ist beim '
                           'Löschen aufgetreten')

    def get_projections_available(self):
        sql = f"""
        SELECT *
        FROM {self.vt_schema}.projections_available
        ORDER BY srid
        """
        return self.fetch(sql)

    def add_projection(self, srid, description):
        sql = f"""
        INSERT INTO {self.vt_schema}.projections_available (srid, description)
        VALUES ('{srid}','{description}');
        """
        return self.execute(sql)

    def get_srid(self, projection_data):
        sql = f"""
        SELECT prjtxt2epsg('{projection_data}')
        """
        return self.fetch(sql)

    def get_spatial_ref(self, srid):
        sql = f"""
        SELECT srtext
        FROM spatial_ref_sys
        WHERE srid = {srid}
        """
        return self.fetch(sql)

    def get_column_names(self, schema, table):
        sql = f"""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = '{schema}'
        AND table_name = '{table}'
        """
        return self.fetch(sql)

    def get_pkey(self, schema, table):
        '''
        get the current primary key of given table
        '''
        sql = f"""
        SELECT a.attname
        FROM   pg_index i
        JOIN   pg_attribute a ON a.attrelid = i.indrelid
                             AND a.attnum = ANY(i.indkey)
        WHERE  i.indrelid = '"{schema}"."{table}"'::regclass
        AND    i.indisprimary;
        """
        ret = [a.attname for a in self.fetch(sql)]
        return ret

    def set_zone(self, schema, table, zone_id_column=None,
                 zone_name_column=''):
        '''
        set the columns with the zone-id, zone-name for the given
        area table makes an entry in the areas_available table, that this column is
        preferred as pkey
        returns True if successful
        returns False if not (e.g. column isn't unique, table doesn't exist)
        '''

        try:
            pkey = self.get_pkey(schema, table)
        except Exception as e:
            return False, e
        if len(pkey) != 1:
            return (False, 'Kein eindeutiger primary key in hochgeladener'
                    'Tabelle vorhanden!')
        pkey = pkey[0]
        if not zone_id_column:
            zone_id_column = pkey
        if not zone_name_column:
            zone_name_column = zone_id_column
        else:
            # set custom zone-id column to unique to check if values are unique
            try:
                # zone-id
                sql_unique = f"""
                ALTER TABLE "{schema}"."{table}"
                ADD UNIQUE ("{zone_id_column}")
                """
                self.execute(sql_unique)
            # catches all exception (e.g. if table doesn't exist)
            except Exception as e:
                return False, e

        sql_update = f"""
        UPDATE {self.vt_schema}.areas_available
        SET pkey='{pkey}', zone_id='{zone_id_column}', zone_name='{zone_name_column}'
        WHERE schema='{schema}'
        AND table_name='{table}'
        """
        # update the meta table
        self.execute(sql_update)
        return True, ''

    def upload_shape(self, schema, name, shapefile, process, conversion_process,
                 on_progress=None, srid=None, on_exit=None, encoding='UTF-8'):
        '''
        upload a shapefile to the database, monitor the progress

        @param schema - the schema the table will be created in
        @param name - the name the area (= the table) gets
        @param shapefile - the path to the shapefile to upload
        @param process - QtCore.QProcess, process provided to upload shape into db
        @param conversion - QtCore.QProcess, process provided to convert file
        @param on_progress - optional, method expecting a string as a parameter
        and an optional progress value (from 0 to 100)
        @param on_exit - optional, additional callback, called when process is done,
        expects exit-code and -status as params
        @param projection - optional, srid of the projection
        @param encoding - optional, encoding of the shapefile
        '''
        psql_path = config.settings['env']['psql_path']
        shp2pgsql_path = config.settings['env']['shp2pgsql_path']

        if not os.path.exists(psql_path):
            psql_path = os.path.join(os.path.dirname(__file__),
                                     psql_path)
        if not os.path.exists(psql_path):
            on_progress(u'<b>Die angegebene <i>psql_path.exe</i> wurde '
                        'nicht gefunden. </br> Bitte prüfen Sie die '
                        'Einstellungen!</b>')
            return

        if not os.path.exists(shp2pgsql_path):
            shp2pgsql_path = os.path.join(os.path.dirname(__file__),
                                          shp2pgsql_path)
        if not os.path.exists(shp2pgsql_path):
            on_progress(u'<b>Die angegebene <i>shp2pgsql_path.exe</i> wurde '
                        'nicht gefunden. </br> Bitte prüfen Sie die '
                        'Einstellungen!</b>')
            return

        options = '-c -g geom -I -i'
        self.on_progress = on_progress
        target_srid = config.settings['db_config']['srid']
        if srid:
            options += ' -s {sourcesrid}:{target_srid}'.format(
                sourcesrid=srid, target_srid=target_srid)
        options += ' -W {}'.format(encoding)
        #put tmp file in folder where this script is located
        tmp_dir = tempfile.mkdtemp()
        tmp_file = os.path.join(tmp_dir, 'temp.sql')
        shp2pgsql_cmd = (f'"{shp2pgsql_path}" {options} "{shapefile}" '
                         f'"{schema}"."{table}"'
                         )

        def finished(exit_code, exit_status):
            on_exit(exit_code, exit_status)
            shutil.rmtree(tmp_dir)

        # call callback with standard error and output
        def progress():
            if process.state() != QtCore.QProcess.NotRunning:
                out = bytearray(process.readAllStandardOutput())
                err = bytearray(process.readAllStandardError())
                if len(out) > 0: on_progress(out.decode('utf-8'))
                if len(err) > 0: on_progress(err.decode('utf-8'))

            if conversion_process.state() != QtCore.QProcess.NotRunning:
                self.on_progress(str(conversion_process.readAllStandardError()))

        if self.on_progress:
            # QProcess emits `readyRead` when there is data to be read
            conversion_process.readyReadStandardError.connect(progress)
            process.readyReadStandardOutput.connect(progress)
            process.readyReadStandardError.connect(progress)

        def upload_to_db(exit_code, exit_status):
            if(exit_code != 0):
                if(self.on_progress):
                    self.on_progress('<b>Konvertierung nicht '
                                     'erfolgreich </b><br>')
                return
            if self.on_progress:
                self.on_progress('<b>Konvertierung erfolgreich. '
                                 'Starte Upload... </b><br>', 50)

            db_config = config.settings['db_config']
            # you can't pass a password to the command-line psql.exe
            # -> set environment variable instead
            sys_env = QtCore.QProcessEnvironment.systemEnvironment()
            sys_env.insert("PGPASSWORD", db_config['password'])
            process.setProcessEnvironment(sys_env)

            psql_cmd = (u'"{executable}" -d {database} -a -h {host} -p {port}'
                        ' -U {user} -w -f "{input_file}"'.format(
                            executable=psql_path, database=db_config['db_name'],
                            host=db_config['host'], port=db_config['port'],
                            user=db_config['username'], input_file=tmp_file))

            process.finished.connect(finished)
            process.start(psql_cmd)

        conversion_process.finished.connect(upload_to_db)

        def read_sql_code():
            txt = conversion_process.readAllStandardOutput()
            with open(tmp_file, 'ab') as openfile:
                openfile.write(txt)

        conversion_process.readyReadStandardOutput.connect(read_sql_code)
        conversion_process.start(shp2pgsql_cmd)

    def add_area(self, schema, name, shapefile, process, conversion_process,
                 on_progress=None, srid=None, on_finish=None, on_success=None,
                 encoding='UTF-8'):
        '''
        add an aggregation area based on a given shapefile to the database, monitor
        the progress

        @param schema - the schema the table will be created in
        @param name - the name the area (= the table) gets
        @param shapefile - the path to the shapefile to upload
        @param process - QtCore.QProcess, process provided to upload shape into db
        @param conversion - QtCore.QProcess, process provided to convert file
        @param on_progress - optional, method expecting a string as a parameter and
        an optional progress value (from 0 to 100)
        @param on_finish - optional, additional callback, called when run is finished
        @param on_success - optional, additional callback, called when run was
        successful
        @param projection - optional, srid of the projection
        '''

        self.on_progress = on_progress

        def on_exit(exit_code, exit_status):
            if exit_code == 0:
                # add new area to areas_available
                sql = f"""
                INSERT INTO {self.vt_schema}.areas_available (area_name, schema, table_name, can_be_deleted)
                VALUES ('{name}','{schema}', '{table}', 'TRUE');
                """
                sql_alter = f"""
                ALTER TABLE "{schema}"."{table}"
                OWNER TO group_osm;
                """
                self.execute(sql)
                self.execute(sql_alter)
                if self.on_progress:
                    self.on_progress('<b>Upload erfolgreich</b><br>')
                self.on_progress = None
            if on_finish:
                on_finish()
            # on_success hast to be called after on_finish (this one is called
            # on error as well)
            if exit_code == 0 and on_success:
                on_success()

        self.upload_shape(schema, name, shapefile, process, conversion_process,
                          on_progress=on_progress, srid=srid, on_exit=on_exit,
                          encoding=encoding)

    def new_intersection(self, schema, table, srid=25832):
        # set functions to scope of following class
        fetch = self.fetch
        execute = self.execute
        vt_schema = self.vt_schema

        class Intersection(QtCore.QThread):
            progress = QtCore.pyqtSignal(str, int)
            error = QtCore.pyqtSignal(str)
            def __init__(self):
                QtCore.QThread.__init__(self)

            def run(self):

                sql_pkey = f"""
                SELECT pkey, zone_name, zone_id
                FROM {vt_schema}.areas_available
                WHERE schema='{schema}'
                AND table_name='{table}';
                """

                sql_queries = f"""
                SELECT * FROM {vt_schema}.queries ORDER BY id;
                """

                try:
                    row = fetch(sql_pkey)[0]
                except IndexError:
                    msg = (f'Tabelle {schema}.{table}.\nnicht in der Datenbank '
                           'vorhanden.')
                    self.error.emit(msg)
                    return

                pkey = row.pkey
                zone_name = row.zone_name
                zone_id = row.zone_id

                queries = fetch(sql_queries)

                if (not pkey) or (not zone_id):
                    msg = (f'Fehlerhafter Eintrag für {schema}.{table}.\nEs '
                           'ist keine zone_id oder kein primary key angegeben.'
                           )
                    self.error.emit(msg)
                    return

                self.progress.emit('Starte Verschneidung...', 0)

                if zone_name is None or zone_name == '':
                    name_str = "''"
                else:
                    name_str = f't."{zone_name}"'
                sql_prep = f"""
                INSERT INTO {vt_schema}.scenario (area_id, start_time, end_time, started, finished)
                SELECT a.id, clock_timestamp(), NULL,  True, False
                FROM {vt_schema}.areas_available AS a
                WHERE a.schema = '{schema}' AND a.table_name = '{table}';
                CREATE OR REPLACE VIEW vz.view_vz_aktuell (vz_id, geom, zone_name, pnt) AS
                SELECT
                t."{zone_id}"::integer AS vz_id,
                st_multi(st_transform(t.geom, {srid}))::geometry(MULTIPOLYGON, {srid}) AS geom,
                {name_str}::text AS zone_name,
                CASE WHEN t.xkoord IS NULL THEN st_pointonsurface(t.geom)::geometry(Point, {srid})
                ELSE st_setsrid(st_makepoint(t.xkoord, t.ykoord), {srid})::geometry(Point, {srid})
                END AS pnt

                FROM "{schema}"."{table}" AS t;
                """
                try:
                    execute(sql_prep)
                except psycopg2.ProgrammingError as e:
                    self.error.emit(str(e))
                    return

                self.add_pnt_column_if_exists(zone_id, name_str)

                weight_sum = sum(q.weight for q in queries)
                progress = 0

                for query in queries:
                    self.progress.emit(query.message, progress)
                    try:
                        execute(query.command)
                    except psycopg2.Error as e:
                        self.error.emit(os.linesep.join((
                            query.command, str(e))))
                        return
                    progress += (query.weight / weight_sum) * 100


                self.progress.emit('Nachbereitungen laufen...', progress)

                sql_post = f"""
                UPDATE {vt_schema}.scenario AS sc SET end_time=clock_timestamp(), started=False, finished=True
                FROM {vt_schema}.areas_available AS a, {vt_schema}.last_area_calculated AS l
                WHERE a.schema='{schema}' AND a.table_name='{table}'
                AND l.id = sc.id;
                """
                execute(sql_post)

                self.progress.emit('Verschneidung beendet!', 100)

                return

            def add_pnt_column_if_exists(self, zone_id, name_str):
                # add point column, if exists
                col_pnt = 'pnt'
                sql_col_exists = f'''
                            SELECT EXISTS (SELECT 1
                            FROM information_schema.columns
                            WHERE table_schema='{schema}'
                            AND table_name='{table}'
                            AND column_name='{col_pnt}');
                            '''
                col_exists = fetch(sql_col_exists)
                if col_exists[0][0]:
                    pnt_col_def = f't."{col_pnt}"::geometry(POINT)'
                else:
                    pnt_col_def = 'st_pointonsurface(t.geom)::geometry(POINT)'

                sql_pnt = f"""
                CREATE OR REPLACE VIEW vz.view_vz_aktuell_pnt (vz_id, pnt, zone_name) AS
                SELECT
                t."{zone_id}"::integer AS vz_id,
                {pnt_col_def} AS pnt,
                {name_str}::text AS zone_name

                FROM "{schema}"."{table}" AS t;"""
                execute(sql_pnt)

        thread = Intersection()
        return thread

    def set_current_scenario(self, scenario: str):
        sql = f"""
        UPDATE {self.vt_schema}.current_scenario
        SET scenario='{scenario}'
        """
        self.execute(sql)

    def get_last_calculated(self):
        sql = f"""
        SELECT *
        FROM {self.vt_schema}.last_area_calculated
        """
        return self.fetch(sql)

    def force_reset_calc(self):
        '''
        resets the database entry for locking calculations to be able to start
        new last row is identified by max id
        warning: if a calculation is still running, new calculation may result
        in errors
        '''
        sql = f"""
        UPDATE {self.vt_schema}.scenario
        SET finished = true
        WHERE id = (SELECT max(id) FROM {self.vt_schema}.scenario )
        """
        self.execute(sql)


    # empty column array selects all columns (*)
    def db_table_to_csv_file(self, schema, table,
                             filename,
                             columns=None,
                             order_by=None):
        order = f' ORDER BY "{order_by}"' if order_by in columns or len(columns) == 0 else ''

        if columns and len(columns) > 0:
            columns=['"{}"'.format(c) for c in columns]
            columns=','.join(columns)
        else:
            columns='*'

        sql = f"""
        COPY (SELECT {columns}
        FROM "{schema}"."{table}"{order}) TO STDOUT WITH CSV HEADER
        """
        with open(filename, 'w') as fileobject:
            self.copy_expert(sql, fileobject)

    def results_to_csv(self,
                       schema: str,
                       table: str,
                       columns: List[str],
                       filename: str):
        columns = ['vz_id', 'zone_name'] + columns
        self.db_table_to_csv_file(schema, table, filename, columns=columns,
                                  order_by='vz_id')

    def results_to_visum_transfer(self,
                                  schema: str,
                                  table: str,
                                  columns: List[str],
                                  filename: str,
                                  visum_classname: str = 'Bezirke',
                                  append: bool = False,
                                  long_format: bool = False):
        columns = ['vz_id'] + columns
        colstr = ', '.join(f'"{c}"' for c in columns)
        sql = f'SELECT {colstr} FROM "{schema}"."{table}"'
        with Connection(self.login) as conn:
            df = pd.read_sql(sql,
                             con=conn,
                             index_col='vz_id')
        save_to_visum_transfer(df, filename, visum_classname, append, long_format)

    def results_to_excel(self,
                         schema: str,
                         table: str,
                         columns: List[str],
                         filename: str,
                         visum_classname: str = 'Bezirke',
                         append: bool = False):
        tmp_dir = tempfile.mkdtemp()
        tmp_filename = os.path.join(tmp_dir, 'temp.csv')
        self.results_to_csv(schema, table, columns, tmp_filename)

        with open(tmp_filename, 'r') as f:
            csvreader = csv.reader((f), delimiter=",")
            if append:
                wbk = load_workbook(filename)
            else:
                wbk = Workbook()
                first_sheet = wbk.active
            sheet = wbk.create_sheet(visum_classname)
            if not append:
                wbk.remove_sheet(first_sheet)
            for r, row in enumerate(csvreader):
                for c, value in enumerate(row):
                    try:
                        val = float(value)
                    except ValueError:
                        val = value
                    sheet.cell(r + 1, c + 1, val)
            wbk.save(filename)

    def db_table_to_shape_file(self, schema, table, process, filename,
                               columns=None,  on_progress=None,
                               srid=None, on_finish=None):
        pgsql2shp_path = config.settings['env']['pgsql2shp_path']
        db_config = config.settings['db_config']

        if not os.path.exists(pgsql2shp_path):
            pgsql2shp_path = os.path.join(os.path.dirname(__file__),
                                          pgsql2shp_path)

        if not os.path.exists(pgsql2shp_path):
            on_progress(u'<b>Die angegebene <i>pgsql2shp.exe</i> wurde nicht '
                        'gefunden. </br> Bitte prüfen Sie die Einstellungen!'
                        '</b>')
            return

        if columns and len(columns) > 0:
            columns = ['"\""{}"\""'.format(c) for c in columns]
            columns = ','.join(columns)
        else:
            columns = '*'

        sql = f'SELECT {columns} FROM "{schema}"."{table}"'

        database = db_config['db_name'],
        host = db_config['host'],
        port = db_config['port'],
        user = db_config['username'],
        password = db_config['password'],

        pgsql2shp_cmd = (f'"{pgsql2shp_path}" -f "{filename}" -h {host} -p {port} '
                         f'-u {user} -P {password} {database} "{sql}"')

        # call callback with standard error and output
        def progress():
            out = bytearray(process.readAllStandardOutput())
            err = bytearray(process.readAllStandardError())
            if len(out) > 0: on_progress(out.decode('utf-8'))
            if len(err) > 0: on_progress(err.decode('utf-8'))

        if on_progress:
            process.readyReadStandardOutput.connect(progress)
            process.readyReadStandardError.connect(progress)

        on_progress(f'Konvertiere {schema}.{table} in {filename}')
        process.start(pgsql2shp_cmd)


    def results_to_shape(self, schema, table, columns, process, filename,
                         on_progress=None, srid=None, on_finish=None):
        columns = ['vz_id', 'zone_name', 'geom'] + columns
        self.db_table_to_shape_file(schema, table, process, filename,
                                   columns=columns, on_progress=on_progress,
                                   srid=srid, on_finish=on_finish)


def parse_projection_file(filename):
    '''
    parse and return the complete data, projcs and geogcs description out of a
    given .prj file
    '''
    with open(filename) as f:
        data = f.read()
        projcs, geogcs = parse_projection_data(data)

    return data, projcs, geogcs

def parse_projection_data(data):
    '''
    parse and return the projcs and geogcs description out of a projection-string
    '''
    projcs_pattern = 'PROJCS\["(.*?)",'
    projcs_matches = re.findall(projcs_pattern, data)
    geogcs_pattern = 'GEOGCS\["(.*?)",'
    geogcs_matches = re.findall(geogcs_pattern, data)

    projcs = projcs_matches[0] if len(projcs_matches) > 0 else None
    geogcs = geogcs_matches[0] if len(geogcs_matches) > 0 else None

    return projcs, geogcs
